import json
import xml.etree.ElementTree as ET
import os
import difflib
import io
import base64
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
import re
from collections import Counter
import posixpath
import requests
import traceback

os.path = posixpath

API_KEY = "AIzaSyCOKPBzAeqOocXuiiH44NGBIGLX-IrnlrY"

def read_file_from_resources(file_path):
    with open(file_path, 'r') as file:
        return file.read()
    
def read_edi_promptes():
    template = read_file_from_resources('/Users/ankit.raj/Developer/mapping_automation_testing/resource/ediPrompt.txt')
    print("Formatted read_edi_promptes:", template)
    return template

def load_main_prompts(source_fields, target_fields):
    template = read_file_from_resources('/Users/ankit.raj/Developer/mapping_automation_testing/resource/mainPrompt.txt')
    
    # Convert lists to comma-separated string or JSON array (as per your use case)
    source_str = ', '.join(source_fields)
    target_str = ', '.join(target_fields)

    # Replace placeholders with actual data
    formatted_template = template.format(sourceNames=source_str, targetNames=target_str)
    print("Formatted main_prompts:", formatted_template)
    return formatted_template

def clean_ai_response(ai_response):
    # Remove Markdown-style code fencing like ```json ... ```
    cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", ai_response.strip(), flags=re.IGNORECASE)
    return cleaned
     

def extract_json_paths(data, parent_key="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []

    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                paths.extend(extract_json_paths(value, new_key, seen_paths))
            else:
                if new_key not in seen_paths:
                    seen_paths.add(new_key)
                    paths.append(new_key)
    elif isinstance(data, list):
        new_key = f"{parent_key}[*]" if parent_key else "[*]"
        if new_key not in seen_paths:
            seen_paths.add(new_key)
            if data:
                paths.extend(extract_json_paths(data[0], new_key, seen_paths))

    return paths

def extract_xml_paths(element, parent_path="", seen_paths=None):
    if seen_paths is None:
        seen_paths = set()
    paths = []

    current_path = f"{parent_path}/{element.tag}" if parent_path else element.tag
    if list(element):
        for child in element:
            paths.extend(extract_xml_paths(child, current_path, seen_paths))
    else:
        if current_path not in seen_paths:
            seen_paths.add(current_path)
            paths.append(current_path)

    return paths

def extract_edifact_fields(content, seg_sep="'", elem_sep="+", sub_elem_sep=":"):
    fields = []
    segment_counts = Counter()
    seen_paths = set()

    segments = content.strip().split(seg_sep)
    for segment in segments:
        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        segment_counts[segment_type] += 1

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue

        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        qualifier = parts[1] if len(parts) > 1 else ""

        for i, element in enumerate(parts[1:], start=1):
            sub_elements = element.split(sub_elem_sep)
            field_index = str(i).zfill(2)

            if len(sub_elements) > 1:
                for j, sub_element in enumerate(sub_elements, start=1):
                    field_path = f"{segment_type}{field_index}.{j}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}.{j}"
                    if field_path not in seen_paths and sub_element.strip():
                        seen_paths.add(field_path)
                        fields.append(field_path)
            else:
                field_path = f"{segment_type}{field_index}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}"
                if field_path not in seen_paths:
                    seen_paths.add(field_path)
                    fields.append(field_path)
    return fields

def extract_x12_fields(content, seg_sep="~", elem_sep="*", sub_elem_sep=":"):
    fields = []
    seen_paths = set()
    segment_counts = Counter()

    segments = content.strip().split(seg_sep)

    for segment in segments:
        parts = segment.strip().split(elem_sep)
        segment_type = parts[0].strip()
        segment_counts[segment_type] += 1

    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue

        parts = segment.split(elem_sep)
        segment_type = parts[0].strip()
        qualifier = parts[1].strip() if len(parts) > 1 else ""

        for i, element in enumerate(parts[1:], start=1):
            sub_elements = element.split(sub_elem_sep)
            field_index = str(i).zfill(2)

            if len(sub_elements) > 1:
                for j, sub_element in enumerate(sub_elements, start=1):
                    field_path = f"{segment_type}{field_index}.{j}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}.{j}"
                    if field_path not in seen_paths and sub_element.strip():
                        seen_paths.add(field_path)
                        fields.append(field_path)
            else:
                field_path = f"{segment_type}{field_index}_{qualifier}" if segment_counts[segment_type] > 1 else f"{segment_type}{field_index}"
                if field_path not in seen_paths:
                    seen_paths.add(field_path)
                    fields.append(field_path)
    return fields

def read_content(content, file_type):
    if file_type == "JSON":
        try:
            json_content = json.loads(content)
            return extract_json_paths(json_content)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON: {e}")
    elif file_type == "XML":
        try:
            return extract_xml_paths(ET.fromstring(content))
        except ET.ParseError as e:
            raise ValueError(f"Invalid XML: {e}")
    elif file_type == "EDIFACT":
        return extract_edifact_fields(content)
    elif file_type == "X12":
        return extract_x12_fields(content)
    else:
        raise ValueError("Unsupported file type")

def call_gemini_api(prompt):
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={API_KEY}"
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.3, "maxOutputTokens": 2048}
        }
        headers = {"Content-Type": "application/json"}
        response = requests.post(url, json=payload, headers=headers)
        response_json = response.json()
        ai_response = response_json["candidates"][0]["content"]["parts"][0]["text"]
        cleaned = clean_ai_response(ai_response)
        return json.loads(cleaned)

    except Exception as e:
        print(f"API Error: {str(e)}")
        return {}

def get_mapping_from_ai(source_fields, target_fields):
    prompt = load_main_prompts(source_fields, target_fields)
    return call_gemini_api(prompt)

def get_edi_field_descriptions(edi_fields):
    edi_prompt = read_edi_promptes()
    prompt = edi_prompt.format(ediFields=json.dumps(edi_fields, indent=2))
    return call_gemini_api(prompt)

def create_excel_mapping(source_fields, target_fields, mappings, edi_descriptions=None):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Field Mapping"

        headers = ["Source Field", "Target Field", "AI Suggested Mapping", "Description"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.font = Font(bold=True)

        for source in source_fields:
            target_match = json.dumps(mappings.get(source, ""), indent=2) if isinstance(mappings.get(source, ""), dict) else mappings.get(source, "")
            description = edi_descriptions.get(source, "") if edi_descriptions else ""
            ws.append([source, "", target_match, description])

        dv = DataValidation(type="list", formula1=f'"{','.join(target_fields)}"', showDropDown=True)
        ws.add_data_validation(dv)
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=1+len(source_fields)):
            for cell in row:
                dv.add(cell)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        return wb
    except Exception as e:
        print(f"Error creating Excel mapping: {str(e)}")
        return None

# Main function
def main(source_type, source_data, target_type, target_data):
    try:
        source_fields = read_content(source_data, source_type)
        target_fields = read_content(target_data, target_type)
        mappings = get_mapping_from_ai(source_fields, target_fields)
        edi_descriptions = None
        if source_type in ["EDIFACT", "X12"] or target_type in ["EDIFACT", "X12"]:
            edi_fields = source_fields if source_type in ["EDIFACT", "X12"] else target_fields
            edi_descriptions = get_edi_field_descriptions(edi_fields)

        wb = create_excel_mapping(source_fields, target_fields, mappings, edi_descriptions)
        file_path = "/Users/ankit.raj/Developer/mapping_automation_testing/FieldMapping.xlsx"
        wb.save(file_path)
        encoded_excel = None
        print(f"Excel file saved at: {file_path}")

        return {
            "status": "success",
            "source_fields": source_fields,
            "target_fields": target_fields,
            "mappings": mappings,
            "edi_descriptions": edi_descriptions,
            "excel_base64": encoded_excel
        }
    except Exception as e:
        tb = traceback.format_exc()
        line_number = None
        for line in tb.splitlines():
            if "File" in line and "line" in line:
                line_number = line.split(",")[1].strip().split(" ")[1]
                break
        return {
            "status": "error",
            "message": str(e),
            "line_number": line_number
        }

# Example use
if __name__ == "__main__":
    source_type = "JSON"
    target_type = "JSON"
    source_data = '{"employee": {"name": "John", "id": 123}}'
    target_data = '{"employee": {"name": "John", "id": 123}}'
    # target_data = 'ISA*00*          *00*          *ZZ*ABCDEFGHIJKLM  *12*123456789012  *210101*1253*U*00401*000000001*0*P*>~GS*PO*4405197800*999999999*20020430*1430*1*X*004010~'
    result = main(source_type, source_data, target_type, target_data)
    print(json.dumps(result, indent=2))